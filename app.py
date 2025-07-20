from flask import Flask, render_template, request, jsonify, send_file, session, send_from_directory
import json, random, csv, os
from PIL import Image, ImageDraw, ImageFont
import time, uuid, hmac, hashlib, base64
import requests
from openpyxl import Workbook, load_workbook  # ✅ Workbook 추가!


# 📁 파일 경로 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PRIZE_FILE = os.path.join(BASE_DIR, 'data', 'prizes.json')
LOG_FILE = os.path.join(BASE_DIR, 'result_log.csv')
CARDS_FOLDER = os.path.join(BASE_DIR, 'static', 'cards')
PRIZES_XLSX_PATH = os.path.join(os.path.dirname(__file__), 'prizes.xlsx')


app = Flask(__name__)
app.secret_key = 'supersecret'

# ✅ 인증번호 저장 딕셔너리
auth_codes = {}

# 📱 네이버 SENS 설정 (필요 시 사용)
SENS_ACCESS_KEY = 'ncp_iam_BPAMKR5Z5aqHdik0H2Ew'
SENS_SECRET_KEY = 'ncp_iam_BPKMKRBZpiXH6GerfRGVD7uAAw9fAg4nAJ'
SENS_SERVICE_ID = 'ncp:sms:kr:352930025196:25_survey'
SENS_SENDER_PHONE = '01053012924'

used_users = set()

def send_sms(phone, code):
    timestamp = str(int(time.time() * 1000))
    uri = f"/sms/v2/services/{SENS_SERVICE_ID}/messages"
    url = f"https://sens.apigw.ntruss.com{uri}"

    message = {
        "type": "SMS",
        "from": SENS_SENDER_PHONE,
        "content": f"[목행용탄동 주민자치회 감사 이벤트] \n [인증번호] {code}",
        "messages": [{"to": phone}]
    }

    access_key = SENS_ACCESS_KEY
    secret_key = bytes(SENS_SECRET_KEY, "utf-8")
    message_str = f"POST {uri}\n{timestamp}\n{access_key}"
    signature = base64.b64encode(
        hmac.new(secret_key, bytes(message_str, "utf-8"), hashlib.sha256).digest()
    ).decode("utf-8")

    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "x-ncp-apigw-timestamp": timestamp,
        "x-ncp-iam-access-key": access_key,
        "x-ncp-apigw-signature-v2": signature,
    }

    res = requests.post(url, headers=headers, json=message)
    return res.status_code == 202  # 202이면 정상 처리

@app.route('/send_code', methods=['POST'])
def send_code():
    data = request.get_json()
    phone = data.get('phone')
    code = str(random.randint(100000, 999999))
    auth_codes[phone] = code

    print('🔔 인증번호 요청 도착:', phone, code, flush=True)
    # return jsonify({'success': True})

    # ✅ 실제 문자 발송
    if send_sms(phone, code):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '문자 발송 실패'})



@app.route('/verify_code', methods=['POST'])
def verify_code():
    data = request.get_json()
    phone = data.get('phone')
    code = data.get('code')

    # ✅ auth_codes에서 직접 확인
    if auth_codes.get(phone) == code:
        session[f'verified_{phone}'] = True
        return jsonify({'success': True})
    return jsonify({'success': False})


@app.route('/check_participation', methods=['POST'])
def check_participation():
    data = request.get_json()
    name = data['name']
    phone = data['phone']
    already = has_already_participated(name, phone)
    return jsonify({'exists': already})


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/winners')
def winners_page():
    return render_template('winners.html')


@app.route('/api/winners')
def get_winner_logs():
    page = int(request.args.get('page', 1))
    per_page = 20
    masked_logs = []

    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, newline='', encoding='utf-8-sig') as f:
            reader = list(csv.reader(f))
            total = len(reader)
            start = (page - 1) * per_page
            end = start + per_page
            for row in reader[start:end]:
                if len(row) >= 3:
                    name, phone, prize = row[:3]
                    name = name[0] + '*' + name[-1] if len(name) >= 3 else name[0] + '*'
                    phone = phone[:3] + '****' + phone[-4:]
                    masked_logs.append({'name': name, 'phone': phone, 'prize': prize})
    return jsonify(masked_logs)


@app.route('/admin')
def admin():
    stats = {}
    logs = []

    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) == 3:
                    name, phone, prize = row
                    serial = '-'
                else:
                    name, phone, prize, serial = row
                logs.append({'name': name, 'phone': phone, 'prize': prize, 'serial': serial})
                stats[prize] = stats.get(prize, 0) + 1

    return render_template('admin.html', stats=stats, logs=logs)


@app.route('/prizes')
def get_prizes():
    with open(PRIZE_FILE, encoding='utf-8') as f:
        prizes = json.load(f)
    return jsonify([{"name": p["name"]} for p in prizes])  # ✅ 반드시 반환


@app.route('/spin', methods=['POST'])
def spin():
    try:
        name = request.form['name']
        phone = request.form['phone']

        if has_already_participated(name, phone):
            return jsonify({'status': 'error', 'message': '이미 참여하셨습니다.'})

        if not session.get(f'verified_{phone}'):
            return jsonify({'status': 'error', 'message': '인증되지 않은 사용자입니다.'})

        # 여기에 중간 로그 추가
        print("✅ 인증된 사용자, 룰렛 진행 시작")

        prizes = load_prizes()
        prize_name = pick_prize(prizes)
        prize_index = next((i for i, p in enumerate(prizes) if p['name'] == prize_name), -1)

        serial_number = get_serial_number(prize_name)
        image_path = os.path.join('static', 'cards', f"{prize_name}.png")  # ✅ 파일 전체 경로 지정
        output_image = add_text_to_image(image_path, name, phone, prize_name, serial_number)


        # prizes.xlsx에 이름/전화번호/시리얼 기록
        if not os.path.exists(PRIZES_XLSX_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = prize_name
            ws.append(['이름', '전화번호', '시리얼'])
        else:
            wb = load_workbook(PRIZES_XLSX_PATH)
            if prize_name in wb.sheetnames:
                ws = wb[prize_name]
            else:
                ws = wb.create_sheet(title=prize_name)
                ws.append(['이름', '전화번호', '시리얼'])

        ws.append([name, phone, serial_number])
        wb.save(PRIZES_XLSX_PATH)

        print("✅ 로그 저장 중:", name, phone, prize_name, serial_number)

        with open(LOG_FILE, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([name, phone, prize_name, serial_number])

        return jsonify({
            'status': 'success',
            'prize': prize_name,
            'index': prize_index,
            'card_image': f"/{output_image}"
        })
    except Exception as e:
        print('🔴 서버 에러 발생:', e)
        return jsonify({'error': str(e)}), 500

@app.route('/download/excel')
def download_excel():
    filepath = PRIZES_XLSX_PATH
    if os.path.exists(filepath):
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='prizes.xlsx'
        )
    return "파일이 존재하지 않습니다.", 404



@app.route('/admin/clear', methods=['POST'])
def clear_data():
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    if os.path.exists(PRIZES_XLSX_PATH):
        os.remove(PRIZES_XLSX_PATH)

    if os.path.exists('prizes.csv'):  # 사용 안 해도 안전 차원
        os.remove('prizes.csv')

    if os.path.exists(CARDS_FOLDER):
        for filename in os.listdir(CARDS_FOLDER):
            name, ext = os.path.splitext(filename)
            if ext.lower() == '.png' and any(char.isdigit() for char in name):
                os.remove(os.path.join(CARDS_FOLDER, filename))

    used_users.clear()
    return jsonify({'status': 'success'})


def load_prizes():
    with open(PRIZE_FILE, encoding='utf-8') as f:
        return json.load(f)


def pick_prize(prizes):
    rand = random.random()
    cumulative = 0.0
    for prize in prizes:
        cumulative += prize['probability']
        if rand < cumulative:
            return prize['name']
    return prizes[-1]['name']


def get_serial_number(prize):
    filename = PRIZES_XLSX_PATH

    # 파일 없으면 생성
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = prize
        ws.append(['이름', '전화번호', '시리얼'])
        wb.save(filename)

    wb = load_workbook(filename)

    # ✅ 이전 시트명이 prize와 다른 경우에도 강제로 prize 이름의 시트 사용
    if prize not in wb.sheetnames:
        ws = wb.create_sheet(title=prize)
        ws.append(['이름', '전화번호', '시리얼'])
    ws = wb[prize]

    # ✅ 현재 시트에서 일련번호 생성
    count = ws.max_row  # 헤더 포함됨
    serial = f"{prize}{str(count).zfill(2)}"

    wb.save(filename)
    return serial




def add_text_to_image(image_path, name, phone, prize, serial_number):
    full_image_path = os.path.join(BASE_DIR, image_path)
    img = Image.open(full_image_path)
    draw = ImageDraw.Draw(img)

    if os.name == 'nt':
        font_path = "C:/Windows/Fonts/malgun.ttf"
    else:
        font_path = "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"

    font = ImageFont.truetype(font_path, 28)

    # 기본 정보 왼쪽 위
    phone_last4 = phone[-4:]
    left_text = f"이름: {name}\n전화번호: {phone_last4}\n상품명: {prize}"
    draw.text((35, 35), left_text, font=font, fill=(0, 0, 0))

    # ✅ 오른쪽 위에 일련번호 출력 (오른쪽 정렬)
    serial_text = f"NO: {serial_number}"
    bbox = draw.textbbox((0, 0), serial_text, font=font)
    text_width = bbox[2] - bbox[0]
    draw.text((img.width - text_width - 35, 35), serial_text, font=font, fill=(0, 0, 0))

    output_filename = f"{serial_number}.png"
    output_path = os.path.join(BASE_DIR, 'static', 'cards', output_filename)
    img.save(output_path)

    return f"static/cards/{output_filename}"

def has_already_participated(name, phone):
    if not os.path.exists(LOG_FILE):
        return False
    with open(LOG_FILE, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2 and row[1] == phone:  # ✅ 전화번호만 비교
                return True
    return False



if __name__ == '__main__':
    app.run(debug=True)
